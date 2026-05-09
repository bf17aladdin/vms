# vms/backend/services/reporting_service.py - Reporting & Automation (Sprint 7)

import logging
import json
from datetime import datetime, timedelta
from typing import List, Dict, Optional
from enum import Enum
from pathlib import Path

logger = logging.getLogger(__name__)

class ReportFormat(str, Enum):
    """Formats d'export"""
    PDF = "pdf"
    EXCEL = "excel"
    JSON = "json"
    CSV = "csv"
    HTML = "html"

class ReportType(str, Enum):
    """Types de rapports"""
    DETECTION_LOG = "detection_log"      # Historique détections
    PERSONNEL_ACTIVITY = "personnel_activity"
    VEHICLE_ACTIVITY = "vehicle_activity"
    ANOMALIES = "anomalies"
    DAILY_SUMMARY = "daily_summary"
    WEEKLY_SUMMARY = "weekly_summary"
    SYSTEM_HEALTH = "system_health"

class ReportingService:
    """
    Service de reporting et d'export.
    Génère rapports / exports / alertes programmées.
    
    **Sprint 7 Deliverable**: Reporting & Automation
    """
    
    def __init__(self, data_dir: str = "data/reports"):
        self.data_dir = Path(data_dir)
        self.data_dir.mkdir(parents=True, exist_ok=True)
        
        self.scheduled_exports: Dict[str, Dict] = {}  # export_id -> config
        self.templates: Dict[ReportType, Dict] = self._load_templates()
    
    def _load_templates(self) -> Dict[ReportType, Dict]:
        """Charges templates de rapports"""
        return {
            ReportType.DETECTION_LOG: {
                "title": "Detection Log Report",
                "columns": ["timestamp", "type", "confidence", "location", "object_id"]
            },
            ReportType.PERSONNEL_ACTIVITY: {
                "title": "Personnel Activity Report",
                "columns": ["date", "personnel_id", "name", "entries", "exits", "duration", "zones"]
            },
            ReportType.VEHICLE_ACTIVITY: {
                "title": "Vehicle Activity Report",
                "columns": ["date", "plate", "entries", "exits", "total_time", "flags"]
            },
            ReportType.ANOMALIES: {
                "title": "Anomaly Report",
                "columns": ["timestamp", "anomaly_type", "severity", "details", "resolution"]
            },
            ReportType.DAILY_SUMMARY: {
                "title": "Daily Summary",
                "columns": ["date", "personnel_count", "vehicle_count", "anomalies", "alerts"]
            },
            ReportType.WEEKLY_SUMMARY: {
                "title": "Weekly Summary",
                "columns": ["week", "total_personnel", "total_vehicles", "anomalies", "trends"]
            },
            ReportType.SYSTEM_HEALTH: {
                "title": "System Health Report",
                "columns": ["timestamp", "cpu_usage", "memory_usage", "cameras_active", "errors"]
            }
        }
    
    def generate_report(
        self,
        report_type: ReportType,
        data: List[Dict],
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> Dict:
        """
        Génère un rapport brut (avant formatting)
        
        Args:
            report_type: Type de rapport
            data: Données brutes
            start_date: Début période
            end_date: Fin période
        
        Returns:
            Dict du rapport avec métadonnées
        """
        template = self.templates.get(report_type, {})
        
        report = {
            'id': f"{report_type.value}_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            'type': report_type.value,
            'template': template.get('title'),
            'generated_at': datetime.now().isoformat(),
            'period': {
                'start': start_date.isoformat() if start_date else None,
                'end': end_date.isoformat() if end_date else None
            },
            'row_count': len(data),
            'columns': template.get('columns', []),
            'data': data
        }
        
        logger.info(f"✓ Report {report['id']} generated with {len(data)} rows")
        return report
    
    def export_report(
        self,
        report: Dict,
        format: ReportFormat,
        filename: Optional[str] = None
    ) -> str:
        """
        Export un rapport dans un format spécifique
        
        Args:
            report: Rapport généré via generate_report()
            format: Format d'export (PDF, EXCEL, JSON, CSV, HTML)
            filename: Nom fichier custom (optionnel)
        
        Returns:
            Chemin du fichier exporté
        """
        if not filename:
            filename = f"{report['id']}.{format.value}"
        
        filepath = self.data_dir / filename
        
        if format == ReportFormat.JSON:
            with open(filepath, 'w') as f:
                json.dump(report, f, indent=2)
        
        elif format == ReportFormat.CSV:
            # CSV export
            import csv
            with open(filepath, 'w', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=report['columns'])
                writer.writeheader()
                writer.writerows(report['data'])
        
        elif format == ReportFormat.EXCEL:
            # EXCEL export (requires openpyxl)
            try:
                import openpyxl
                from openpyxl.utils import get_column_letter
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = report['type']
                
                # Headers
                for col, header in enumerate(report['columns'], 1):
                    ws.cell(1, col, header)
                
                # Data
                for row, row_data in enumerate(report['data'], 2):
                    for col, header in enumerate(report['columns'], 1):
                        ws.cell(row, col, row_data.get(header))
                
                wb.save(filepath)
            except ImportError:
                logger.warning("openpyxl not installed, falling back to JSON")
                return self.export_report(report, ReportFormat.JSON, filename)
        
        elif format == ReportFormat.PDF:
            # PDF export (requires reportlab)
            try:
                from reportlab.lib.pagesizes import letter
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
                from reportlab.lib import colors
                
                doc = SimpleDocTemplate(str(filepath), pagesize=letter)
                elements = []
                
                # Title
                styles = getSampleStyleSheet()
                title = Paragraph(report['template'], styles['Title'])
                elements.append(title)
                
                # Data table
                table_data = [report['columns']]
                for row in report['data']:
                    table_data.append([str(row.get(col, '')) for col in report['columns']])
                
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table)
                
                doc.build(elements)
            except ImportError:
                logger.warning("reportlab not installed, falling back to HTML")
                return self.export_report(report, ReportFormat.HTML, filename)
        
        elif format == ReportFormat.HTML:
            # HTML export - build table without nested f-strings
            headers_html = ''.join(f'<th>{col}</th>' for col in report['columns'])
            
            rows_html = []
            for row in report['data']:
                cells = ''.join(f'<td>{row.get(col, "")}</td>' for col in report['columns'])
                rows_html.append(f'<tr>{cells}</tr>')
            rows_table = '\n'.join(rows_html)
            
            html = f"""<html>
<head>
    <title>{report['template']}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        table {{ border-collapse: collapse; width: 100%; }}
        th, td {{ border: 1px solid black; padding: 8px; text-align: left; }}
        th {{ background-color: #4CAF50; color: white; }}
    </style>
</head>
<body>
    <h1>{report['template']}</h1>
    <p>Generated: {report['generated_at']}</p>
    <table>
        <tr>{headers_html}</tr>
        {rows_table}
    </table>
</body>
</html>"""
            with open(filepath, 'w') as f:
                f.write(html)
        
        logger.info(f"✓ Report exported to {filepath}")
        return str(filepath)
    
    def schedule_export(
        self,
        report_type: ReportType,
        format: ReportFormat,
        frequency: str,  # "daily", "weekly", "monthly"
        time_of_day: str = "00:00"  # HH:MM
    ) -> str:
        """
        Planifie un export automatique (nécessite scheduler)
        
        Args:
            report_type: Type de rapport
            format: Format d'export
            frequency: Fréquence ("daily", "weekly", "monthly")
            time_of_day: Heure d'exécution (HH:MM)
        
        Returns:
            ID du job planifié
        """
        job_id = f"export_{report_type.value}_{frequency}"
        
        self.scheduled_exports[job_id] = {
            'report_type': report_type.value,
            'format': format.value,
            'frequency': frequency,
            'time_of_day': time_of_day,
            'created_at': datetime.now().isoformat(),
            'last_run': None
        }
        
        logger.info(f"✓ Export job {job_id} scheduled: {frequency} @ {time_of_day}")
        return job_id
    
    def cancel_export(self, job_id: str) -> bool:
        """Annule un export programmé"""
        if job_id in self.scheduled_exports:
            del self.scheduled_exports[job_id]
            logger.info(f"✓ Export job {job_id} cancelled")
            return True
        return False
    
    def get_scheduled_exports(self) -> Dict:
        """Liste tous les exports programmés"""
        return self.scheduled_exports.copy()
    
    def save_scheduled_exports(self):
        """Persiste exports programmés en JSON"""
        config_path = self.data_dir / "scheduled_exports.json"
        with open(config_path, 'w') as f:
            json.dump(self.scheduled_exports, f, indent=2)
        logger.info(f"✓ Scheduled exports saved to {config_path}")

# Instance globale
_reporting: Optional[ReportingService] = None

def get_reporting_service() -> ReportingService:
    """Get or create global reporting service"""
    global _reporting
    if _reporting is None:
        _reporting = ReportingService()
    return _reporting
